#!/usr/bin/env python3
"""
Reemplaza los eventos del calendario médico de agosto a diciembre de 2026.

La fuente de importación es PROGRAMACIÓN DETALLADA dentro del Excel. Antes de
borrar, guarda un respaldo local completo de events y event_participants del
periodo objetivo. Todas las citas se validan antes de modificar Supabase.
"""

from __future__ import annotations

import json
import sys
import unicodedata
import uuid
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import requests


ROOT = Path("/Users/javierescobedo/athlete-platform")
EXCEL_FILE = Path(
    "/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/"
    "CALENDARIO_CITAS_CIRCUITO AGOST-DIC_130826.xlsx"
)
ENV_FILE = ROOT / "apps/web/.env.local"
BACKUP_FILE = ROOT / "scripts/backups/calendar_aug_dec_2026_pre_replace.json"

PERIOD_START = "2026-08-01T00:00:00Z"
PERIOD_END = "2027-01-01T00:00:00Z"
MEXICO_TZ = timezone(timedelta(hours=-6))
BATCH_SIZE = 200

# El Excel tiene "Narváe"; el registro existente en Supabase es "Narváez".
MANUAL_ATHLETE_NAMES = {
    47: "Jimena Espinoza Narváez",
}

EVENT_TYPE_BY_SERVICE = {
    "MÉDICO": "medical",
    "NUTRICIÓN": "nutrition",
    "PSICOLOGÍA": "psychology",
    "FISIOTERAPIA": "physio",
}

MINUTES_BY_SERVICE = {
    "MÉDICO": 20,
    "NUTRICIÓN": 20,
    "PSICOLOGÍA": 20,
    "FISIOTERAPIA": 30,
}


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if "=" not in line or line.lstrip().startswith("#"):
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def normalize(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    return " ".join(
        decomposed.encode("ascii", "ignore").decode("ascii").lower().split()
    )


class SupabaseRest:
    def __init__(self, url: str, service_key: str) -> None:
        self.url = url.rstrip("/")
        self.headers = {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }

    def fetch_all(self, table: str, params: dict[str, str]) -> list[dict]:
        records: list[dict] = []
        offset = 0
        while True:
            response = requests.get(
                f"{self.url}/rest/v1/{table}",
                headers=self.headers,
                params={**params, "limit": 1000, "offset": offset},
                timeout=60,
            )
            response.raise_for_status()
            page = response.json()
            records.extend(page)
            if len(page) < 1000:
                return records
            offset += 1000

    def insert(self, table: str, records: list[dict]) -> None:
        response = requests.post(
            f"{self.url}/rest/v1/{table}",
            headers={**self.headers, "Prefer": "return=minimal"},
            data=json.dumps(records),
            timeout=60,
        )
        response.raise_for_status()

    def delete_period_events(self) -> int:
        """Delete calendar events in scope; participant rows cascade automatically."""
        response = requests.delete(
            f"{self.url}/rest/v1/events",
            headers={**self.headers, "Prefer": "return=representation"},
            params={
                "and": (
                    f"(start_at.gte.{PERIOD_START},"
                    f"start_at.lt.{PERIOD_END})"
                ),
                "select": "id",
            },
            timeout=120,
        )
        response.raise_for_status()
        return len(response.json())

    def exact_count(self, table: str, params: dict[str, str]) -> int:
        response = requests.get(
            f"{self.url}/rest/v1/{table}",
            headers={**self.headers, "Prefer": "count=exact"},
            params={**params, "select": "id", "limit": 1},
            timeout=60,
        )
        response.raise_for_status()
        content_range = response.headers.get("Content-Range", "")
        count = content_range.rsplit("/", 1)[-1]
        return int(count) if count.isdigit() else len(response.json())


def parse_workbook() -> tuple[list[dict], dict[int, str]]:
    workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    assigned_names: dict[int, str] = {}
    for row in workbook["ASIGNACIÓN FIJA"].iter_rows(min_row=2, values_only=True):
        folio, name = row[0], row[1]
        if folio is not None and name is not None:
            assigned_names[int(folio)] = str(name).strip()

    appointments: list[dict] = []
    for row in workbook["PROGRAMACIÓN DETALLADA"].iter_rows(
        min_row=2, values_only=True
    ):
        (
            month,
            week,
            date_text,
            service,
            professional,
            time_text,
            folio,
            name,
            discipline,
            group,
        ) = row
        if folio is None:
            continue
        service = str(service).strip()
        if service not in EVENT_TYPE_BY_SERVICE:
            raise ValueError(f"Servicio no soportado: {service!r}")
        if int(folio) not in assigned_names:
            raise ValueError(f"Folio {folio} no existe en ASIGNACIÓN FIJA")
        if str(name).strip() != assigned_names[int(folio)]:
            raise ValueError(f"Nombre inconsistente en folio {folio}")
        appointments.append(
            {
                "month": str(month).strip(),
                "week": str(week).strip(),
                "date_text": str(date_text).strip(),
                "service": service,
                "professional": str(professional).strip(),
                "time_text": str(time_text).strip(),
                "folio": int(folio),
                "name": str(name).strip(),
                "discipline": str(discipline).strip(),
                "group": group,
            }
        )

    if len(appointments) != 2320:
        raise ValueError(f"Se esperaban 2,320 citas; se encontraron {len(appointments)}")
    return appointments, assigned_names


def build_import_records(
    appointments: list[dict],
    athletes_by_name: dict[str, str],
    created_by_profile_id: str,
) -> tuple[list[dict], list[dict]]:
    events: list[dict] = []
    participants: list[dict] = []
    unique_keys: set[tuple] = set()
    unresolved: list[dict] = []

    for appointment in appointments:
        athlete_name = MANUAL_ATHLETE_NAMES.get(
            appointment["folio"], appointment["name"]
        )
        athlete_id = athletes_by_name.get(normalize(athlete_name))
        if not athlete_id:
            unresolved.append(appointment)
            continue

        start_at = datetime.strptime(
            f"{appointment['date_text']} {appointment['time_text']}",
            "%d/%m/%Y %H:%M",
        ).replace(tzinfo=MEXICO_TZ)
        end_at = start_at + timedelta(minutes=MINUTES_BY_SERVICE[appointment["service"]])
        uniqueness_key = (
            athlete_id,
            appointment["service"],
            appointment["professional"],
            start_at.isoformat(),
        )
        if uniqueness_key in unique_keys:
            raise ValueError(f"Cita duplicada en Excel: {uniqueness_key}")
        unique_keys.add(uniqueness_key)

        event_id = str(uuid.uuid4())
        events.append(
            {
                "id": event_id,
                # Conserva los títulos por equipo que usa el calendario actual.
                "title": appointment["professional"],
                "event_type": EVENT_TYPE_BY_SERVICE[appointment["service"]],
                "start_at": start_at.isoformat(),
                "end_at": end_at.isoformat(),
                "status": "scheduled",
                "created_by_profile_id": created_by_profile_id,
            }
        )
        participants.append(
            {
                "event_id": event_id,
                "participant_id": athlete_id,
                "participant_type": "athlete",
                "attendance_status": "planned",
            }
        )

    if unresolved:
        names = sorted({f"{a['folio']}: {a['name']}" for a in unresolved})
        raise ValueError(f"Atletas sin mapeo: {names}")
    return events, participants


def batch_insert(client: SupabaseRest, table: str, records: list[dict]) -> None:
    for offset in range(0, len(records), BATCH_SIZE):
        batch = records[offset : offset + BATCH_SIZE]
        client.insert(table, batch)
        print(f"  {table}: {offset + len(batch)}/{len(records)}", flush=True)


def main() -> None:
    if "--execute" not in sys.argv:
        raise SystemExit("Ejecuta con --execute después de la validación.")

    env = load_env(ENV_FILE)
    client = SupabaseRest(
        env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]
    )

    print("Validando programación detallada...")
    appointments, _ = parse_workbook()
    print(f"  ✓ {len(appointments)} citas válidas")

    athletes = client.fetch_all(
        "athletes", {"select": "id,first_name,last_name,status"}
    )
    athletes_by_name = {
        normalize(f"{athlete['first_name']} {athlete['last_name']}"): athlete["id"]
        for athlete in athletes
    }

    profiles = client.fetch_all("profiles", {"select": "id", "order": "created_at"})
    if not profiles:
        raise ValueError("No hay profiles disponibles para created_by_profile_id")
    created_by_profile_id = profiles[0]["id"]

    events, participants = build_import_records(
        appointments, athletes_by_name, created_by_profile_id
    )
    print(
        f"  ✓ {len(events)} eventos y {len(participants)} participantes listos "
        "(incluye las 28 citas de fisioterapia del 02/12)"
    )

    print("Respaldando calendario actual...")
    current_events = client.fetch_all(
        "events",
        {
            "select": "*",
            "and": f"(start_at.gte.{PERIOD_START},start_at.lt.{PERIOD_END})",
        },
    )
    event_ids = [event["id"] for event in current_events]
    current_participants: list[dict] = []
    for offset in range(0, len(event_ids), BATCH_SIZE):
        ids = ",".join(event_ids[offset : offset + BATCH_SIZE])
        current_participants.extend(
            client.fetch_all(
                "event_participants",
                {"select": "*", "event_id": f"in.({ids})"},
            )
        )
    BACKUP_FILE.parent.mkdir(parents=True, exist_ok=True)
    BACKUP_FILE.write_text(
        json.dumps(
            {
                "period_start": PERIOD_START,
                "period_end": PERIOD_END,
                "events": current_events,
                "event_participants": current_participants,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(
        f"  ✓ respaldo: {len(current_events)} eventos y "
        f"{len(current_participants)} participantes"
    )

    print("Borrando calendario anterior de agosto a diciembre...")
    deleted = client.delete_period_events()
    if deleted != len(current_events):
        raise RuntimeError(
            f"Borrado incompleto: se esperaban {len(current_events)} y se borraron {deleted}"
        )
    print(f"  ✓ {deleted} eventos eliminados; participantes eliminados por cascada")

    print("Importando eventos...")
    batch_insert(client, "events", events)

    print("Importando participantes...")
    batch_insert(client, "event_participants", participants)

    print("Verificando importación...")
    event_count = client.exact_count(
        "events",
        {"and": f"(start_at.gte.{PERIOD_START},start_at.lt.{PERIOD_END})"},
    )
    imported_event_ids = [
        event["id"]
        for event in client.fetch_all(
            "events",
            {
                "select": "id",
                "and": f"(start_at.gte.{PERIOD_START},start_at.lt.{PERIOD_END})",
            },
        )
    ]
    participant_count = 0
    for offset in range(0, len(imported_event_ids), BATCH_SIZE):
        ids = ",".join(imported_event_ids[offset : offset + BATCH_SIZE])
        participant_count += len(
            client.fetch_all(
                "event_participants",
                {"select": "id", "event_id": f"in.({ids})"},
            )
        )
    if event_count != 2320 or participant_count != 2320:
        raise RuntimeError(
            f"Verificación fallida: events={event_count}, participants={participant_count}"
        )

    print("✓ Sustitución completada: 2,320 eventos y 2,320 participantes.")
    print(f"✓ Respaldo disponible en: {BACKUP_FILE}")


if __name__ == "__main__":
    main()
