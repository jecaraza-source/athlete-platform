"""
sync_training_plans.py
----------------------
Lee todos los Excels de planes de entrenamiento, compara con la BD y:
  1. Inserta datos del Excel si el atleta no tiene evaluación coach
  2. Rellena campos vacíos (NULL) con "pendiente" para todos los atletas
  3. Imprime un resumen de lo que faltaba y lo que se actualizó

Uso:
  SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... python3 scripts/sync_training_plans.py [--dry-run]
"""
import os, re, sys, unicodedata, ssl, json
import urllib.request, urllib.parse
import openpyxl

DRY_RUN = '--dry-run' in sys.argv

SUPABASE_URL = os.environ['SUPABASE_URL']
SUPABASE_KEY = os.environ['SUPABASE_SERVICE_ROLE_KEY']

SSL_CTX = ssl._create_unverified_context()

FILES = [
    '/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE Atletismo.xlsx',
    '/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE Breaking.xlsx',
    '/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE Canotaje.xlsx',
    '/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE GAF.xlsx',
    '/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE Natacion.xlsx',
    '/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE TKD Dos.xlsx',
    '/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE TKD Uno.xlsx',
]

# Row → DB field mapping (col C = column index 3 in openpyxl, 1-indexed)
ROW_TO_FIELD = {
    10: 'strength_test',
    13: 'power_test',
    16: 'speed_test',
    19: 'endurance_test',
    22: 'flexibility_test',
    26: 'technical_weaknesses',
    29: 'competitive_capabilities',
    33: 'movement_efficiency',
    36: 'body_mechanics',
    39: 'segment_alignment',
    43: 'athlete_sport_profile',
    50: 'discipline_intervention',
    54: 'season_structure',
    57: 'competitive_calendar',
    60: 'performance_objectives',
    63: 'preparation_stages',
    67: 'technical_correction',
    70: 'load_supervision',
    73: 'competition_preparation',
    76: 'performance_analysis',
    79: 'continuous_feedback',
    82: 'mark_monitoring',
    85: 'plan_adjustments',
    88: 'observations',
}

ALL_FIELDS = list(ROW_TO_FIELD.values())
SECTION_KEYS = ['medico', 'nutricion', 'psicologia', 'entrenador', 'fisioterapia']

# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def api_get(path):
    req = urllib.request.Request(
        SUPABASE_URL + path,
        headers={'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}
    )
    return json.loads(urllib.request.urlopen(req, context=SSL_CTX).read())

def api_post(path, data):
    body = json.dumps(data).encode()
    req = urllib.request.Request(
        SUPABASE_URL + path,
        data=body,
        headers={
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'return=representation',
        },
        method='POST'
    )
    return json.loads(urllib.request.urlopen(req, context=SSL_CTX).read())

def api_patch(path, data):
    body = json.dumps(data).encode()
    req = urllib.request.Request(
        SUPABASE_URL + path,
        data=body,
        headers={
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'return=minimal',
        },
        method='PATCH'
    )
    urllib.request.urlopen(req, context=SSL_CTX).read()

def api_upsert(path, data, on_conflict):
    body = json.dumps(data).encode()
    url = SUPABASE_URL + path
    if on_conflict:
        url += f'?on_conflict={on_conflict}'
    req = urllib.request.Request(
        url,
        data=body,
        headers={
            'apikey': SUPABASE_KEY,
            'Authorization': f'Bearer {SUPABASE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'resolution=merge-duplicates,return=representation',
        },
        method='POST'
    )
    try:
        return json.loads(urllib.request.urlopen(req, context=SSL_CTX).read())
    except urllib.error.HTTPError as e:
        # 409 = already exists with same section_id; fall back to PATCH
        if e.code == 409:
            athlete_id = data.get('athlete_id')
            if athlete_id:
                patch_data = {k: v for k, v in data.items()
                              if k not in ('athlete_id', 'diagnostic_section_id')}
                api_patch(f'/rest/v1/athlete_coach_evaluation?athlete_id=eq.{athlete_id}',
                          patch_data)
                return []
        raise

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize(text):
    text = str(text).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', text)

def fuzzy_match(db_athletes, full_name):
    needle = set(normalize(full_name).split())
    best, best_score = None, 0.0
    for a in db_athletes:
        hay = set(normalize(f"{a['first_name']} {a['last_name']}").split())
        common = needle & hay
        score = len(common) / max(len(needle), len(hay))
        if score > best_score:
            best_score, best = score, a
    return best if best_score >= 0.4 else None

def parse_sheet(ws):
    """Extract athlete name (row 3, col C) and all training fields."""
    def cell(row, col=3):
        v = ws.cell(row=row, column=col).value
        if v is None: return None
        s = str(v).strip()
        return s if s else None

    fields = {}
    for row_idx, field_name in ROW_TO_FIELD.items():
        fields[field_name] = cell(row_idx)

    return {
        'full_name': cell(3),
        'fields': fields,
    }

def ensure_diagnostic_and_section(athlete_id):
    """Ensure athlete has diagnostic + all 5 sections. Return entrenador section."""
    res = api_get(f'/rest/v1/athlete_initial_diagnostic?select=id&athlete_id=eq.{athlete_id}')
    if res:
        diag_id = res[0]['id']
    else:
        ins = api_post('/rest/v1/athlete_initial_diagnostic',
                       {'athlete_id': athlete_id})
        diag_id = ins[0]['id']
        api_post('/rest/v1/athlete_diagnostic_sections',
                 [{'diagnostic_id': diag_id, 'athlete_id': athlete_id, 'section': s}
                  for s in SECTION_KEYS])

    sec = api_get(f'/rest/v1/athlete_diagnostic_sections?select=id,diagnostic_id'
                  f'&athlete_id=eq.{athlete_id}&section=eq.entrenador')
    if sec:
        return sec[0]

    new_sec = api_post('/rest/v1/athlete_diagnostic_sections',
                       {'diagnostic_id': diag_id, 'athlete_id': athlete_id, 'section': 'entrenador'})
    return new_sec[0]

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"{'[DRY RUN] ' if DRY_RUN else ''}Sincronizando planes de entrenamiento...\n")

    # Load all athletes from DB
    db_athletes = api_get('/rest/v1/athletes?select=id,first_name,last_name,discipline&status=eq.active&order=last_name')
    print(f"Atletas activos en BD: {len(db_athletes)}\n")

    # Load existing coach evaluations
    existing_evals = api_get('/rest/v1/athlete_coach_evaluation?select=athlete_id,diagnostic_section_id,' + ','.join(ALL_FIELDS))
    eval_by_athlete = {e['athlete_id']: e for e in existing_evals}

    # Stats
    stats = {
        'excel_athletes': 0,
        'matched': 0,
        'not_matched': [],
        'inserted': 0,
        'fields_updated': 0,
        'already_complete': 0,
        'by_discipline': {},
    }

    for filepath in FILES:
        discipline = os.path.basename(filepath).replace('PE ', '').replace('.xlsx', '')
        print(f"━━━ {discipline} ({'TKD' if 'TKD' in discipline else discipline}) ━━━")

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"  ERROR al leer: {e}")
            continue

        disc_stats = {'matched': 0, 'missing': 0, 'fields_set': 0, 'not_found': [], 'already_complete': 0}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parsed = parse_sheet(ws)
            full_name = parsed['full_name']
            if not full_name:
                continue

            stats['excel_athletes'] += 1
            athlete = fuzzy_match(db_athletes, full_name)

            if not athlete:
                disc_stats['not_found'].append(full_name)
                stats['not_matched'].append(f"{full_name} ({discipline})")
                continue

            disc_stats['matched'] += 1
            stats['matched'] += 1
            athlete_id = athlete['id']
            excel_fields = parsed['fields']

            existing = eval_by_athlete.get(athlete_id)

            if not existing:
                # Create evaluation with Excel data, using 'pendiente' for null Excel fields
                disc_stats['missing'] += 1
                if not DRY_RUN:
                    section = ensure_diagnostic_and_section(athlete_id)
                    payload = {
                        'diagnostic_section_id': section['id'],
                        'athlete_id': athlete_id,
                    }
                    for field, value in excel_fields.items():
                        payload[field] = value if value else 'pendiente'
                    api_upsert('/rest/v1/athlete_coach_evaluation', payload, 'diagnostic_section_id')
                    disc_stats['fields_set'] += len(ALL_FIELDS)
                    stats['inserted'] += 1
                print(f"  [NUEVO]    {full_name} → {athlete['first_name']} {athlete['last_name']}")
            else:
                # Fill only NULL fields (don't overwrite existing data)
                null_fields = {}
                for field in ALL_FIELDS:
                    if not existing.get(field):  # NULL or empty
                        excel_val = excel_fields.get(field)
                        null_fields[field] = excel_val if excel_val else 'pendiente'

                if null_fields:
                    if not DRY_RUN:
                        api_patch(f'/rest/v1/athlete_coach_evaluation?athlete_id=eq.{athlete_id}',
                                  null_fields)
                    disc_stats['fields_set'] += len(null_fields)
                    stats['fields_updated'] += len(null_fields)
                    fields_list = ', '.join(null_fields.keys())
                    print(f"  [UPDATE]   {athlete['first_name']} {athlete['last_name']} — {len(null_fields)} campos: {fields_list[:80]}")
                else:
                    disc_stats['already_complete'] += 1

        print(f"  Coincidencias: {disc_stats['matched']}/{disc_stats['matched']+len(disc_stats['not_found'])}")
        if disc_stats['not_found']:
            print(f"  Sin match:    {', '.join(disc_stats['not_found'])}")
        print(f"  Campos escritos: {disc_stats['fields_set']}")
        print()

        stats['by_discipline'][discipline] = disc_stats

    # Summary
    print("=" * 60)
    print("RESUMEN FINAL")
    print("=" * 60)
    print(f"  Atletas en Excel:           {stats['excel_athletes']}")
    print(f"  Emparejados con BD:         {stats['matched']}")
    print(f"  Sin match (no encontrados): {len(stats['not_matched'])}")
    print(f"  Evaluaciones nuevas:        {stats['inserted']}")
    print(f"  Campos actualizados:        {stats['fields_updated']}")
    if stats['not_matched']:
        print(f"\n  ATLETAS NO ENCONTRADOS EN BD:")
        for n in stats['not_matched']:
            print(f"    - {n}")
    if DRY_RUN:
        print("\n[DRY RUN] No se realizaron cambios. Quita --dry-run para aplicar.")

if __name__ == '__main__':
    main()
