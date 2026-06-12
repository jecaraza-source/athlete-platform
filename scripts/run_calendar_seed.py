#!/usr/bin/env python3
"""
run_calendar_seed.py
====================
Inserta las citas del calendario directamente en Supabase vía REST API.

Uso:
    python3 scripts/run_calendar_seed.py

Requiere:
    - openpyxl   (ya instalado)
    - requests   (ya instalado)
    - .env.local con NEXT_PUBLIC_SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY
"""

import re
import json
import unicodedata
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests
import openpyxl

# ──────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────────
EXCEL_FILE = (
    "/Users/javierescobedo/Library/Containers/net.whatsapp.WhatsApp"
    "/Data/tmp/documents/84D311DE-F91F-4798-B645-0E33CC12C633"
    "/CALENDARIO CITAS JUNIO-DICIEMBRE_ATLETAS_70626.xlsx"
)
ENV_FILE = "/Users/javierescobedo/athlete-platform/apps/web/.env.local"

SLOT_MINUTES = 30
TZ = "-06:00"
BATCH_SIZE = 200   # registros por llamada a la API


# ──────────────────────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────────────────────
def load_env(path: str) -> dict:
    env = {}
    with open(path) as f:
        for line in f:
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def normalize(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    return " ".join(nfkd.encode("ascii", "ignore").decode("ascii").lower().split())


def add_minutes(time_str: str, minutes: int) -> str:
    h, m = map(int, time_str.split(":"))
    dt = datetime(2000, 1, 1, h, m) + timedelta(minutes=minutes)
    return dt.strftime("%H:%M")


def build_ts(date_str: str, time_str: str) -> str:
    return f"{date_str}T{time_str}:00{TZ}"


def service_to_event_type(service: str) -> str:
    # Valid values: training, competition, meeting, medical, evaluation, other
    # All calendar appointments are medical/clinical → 'medical'
    return "medical"


# ──────────────────────────────────────────────────────────────
# SUPABASE HELPER
# ──────────────────────────────────────────────────────────────
class SupabaseClient:
    def __init__(self, url: str, service_key: str):
        self.url = url.rstrip("/")
        self.headers = {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
        }

    def select(self, table: str, select: str = "*", extra_params: dict | None = None) -> list:
        """Descarga TODOS los registros paginando de 1000 en 1000."""
        params = {"select": select, "limit": 1000}
        if extra_params:
            params.update(extra_params)
        all_rows = []
        offset = 0
        while True:
            params["offset"] = offset
            resp = requests.get(
                f"{self.url}/rest/v1/{table}",
                headers={**self.headers, "Accept": "application/json"},
                params=params,
            )
            resp.raise_for_status()
            chunk = resp.json()
            all_rows.extend(chunk)
            if len(chunk) < 1000:
                break
            offset += 1000
        return all_rows

    def count(self, table: str, extra_params: dict | None = None) -> int:
        """Devuelve el conteo exacto de filas usando el header Content-Range."""
        params = {"select": "id", "limit": 1}
        if extra_params:
            params.update(extra_params)
        resp = requests.get(
            f"{self.url}/rest/v1/{table}",
            headers={**self.headers, "Accept": "application/json",
                     "Prefer": "count=exact"},
            params=params,
        )
        resp.raise_for_status()
        # Content-Range: 0-0/2972
        cr = resp.headers.get("Content-Range", "")
        total = cr.split("/")[-1] if "/" in cr else "?"
        return int(total) if total.isdigit() else len(resp.json())

    def insert_batch(self, table: str, records: list) -> list:
        """Inserta un lote y devuelve las filas insertadas (con id)."""
        resp = requests.post(
            f"{self.url}/rest/v1/{table}",
            headers={**self.headers, "Prefer": "return=representation"},
            data=json.dumps(records),
        )
        if resp.status_code not in (200, 201):
            print(f"  ERROR {resp.status_code}: {resp.text[:300]}")
            resp.raise_for_status()
        return resp.json()


# ──────────────────────────────────────────────────────────────
# PARSER EXCEL  (igual que seed_calendar.py)
# ──────────────────────────────────────────────────────────────
def parse_base_de_datos(wb) -> dict:
    ws = wb["Base de Datos"]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        folio, nombre = row[0], row[1]
        if folio is None or nombre is None:
            continue
        try:
            folio = int(folio)
        except (ValueError, TypeError):
            continue
        mapping[folio] = str(nombre).strip()
    return mapping


def parse_informacion(wb) -> list:
    ws = wb["Informacion"]
    appointments = []
    current_year = 2026
    date_cols: dict = {}
    current_service = None
    in_data = False
    skip_next = False

    for row in ws.iter_rows(values_only=True):
        row = list(row)

        def cell(i):
            v = row[i] if i < len(row) else None
            return str(v).strip() if v is not None else ""

        col0 = cell(0)
        col2 = cell(2)

        if "CALENDARIO DE CITAS" in col0:
            m = re.search(r"(\d{4})", col0)
            current_year = int(m.group(1)) if m else current_year
            in_data, date_cols, current_service, skip_next = False, {}, None, False
            continue

        if col0 == "SERVICIO":
            date_cols = {
                ci: str(val).strip()
                for ci, val in enumerate(row)
                if val is not None and re.match(r"^\d{2}/\d{2}$", str(val).strip())
            }
            in_data, current_service, skip_next = True, None, True
            continue

        if skip_next:
            skip_next = False
            continue

        if "TOTAL DE CITAS" in col0:
            in_data = False
            continue

        if not in_data or not date_cols:
            continue

        if col0 and col0 not in ("", "None"):
            current_service = col0

        if not re.match(r"^\d{2}:\d{2}$", col2):
            continue

        for ci, date_raw in date_cols.items():
            cell_val = row[ci] if ci < len(row) else None
            if cell_val is None:
                continue
            try:
                folio = int(cell_val)
            except (ValueError, TypeError):
                continue
            day, month = date_raw.split("/")
            appointments.append({
                "service":  current_service,
                "date":     f"{current_year}-{month.zfill(2)}-{day.zfill(2)}",
                "time":     col2,
                "folio":    folio,
            })

    return appointments


# ──────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────
def main():
    # 1. Cargar credenciales
    print("Cargando credenciales …")
    env = load_env(ENV_FILE)
    sb = SupabaseClient(env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"])

    # 2. Obtener un profile ID para created_by_profile_id
    print("Obteniendo profile de administrador …")
    profiles = sb.select("profiles", select="id,email")
    if not profiles:
        print("ERROR: no se encontraron profiles en la BD.")
        return
    # Usar el primer profile como autor de las citas
    admin_profile_id = profiles[0]["id"]
    print(f"  → Usando profile: {profiles[0].get('email', admin_profile_id)}")

    # 3. Obtener atletas de la BD
    print("Obteniendo atletas de Supabase …")
    athletes_db = sb.select("athletes", select="id,first_name,last_name")
    name_to_id: dict[str, str] = {}
    for a in athletes_db:
        full = f"{a['first_name']} {a['last_name']}"
        name_to_id[normalize(full)] = a["id"]
    print(f"  → {len(name_to_id)} atletas en la BD")

    # 3. Parsear Excel
    print("Parseando Excel …")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    folio_map = parse_base_de_datos(wb)
    appointments = parse_informacion(wb)
    print(f"  → {len(appointments)} citas en el Excel")

    # 5. Resolver athlete_id para cada cita
    resolved, unmatched_map = [], {}
    for appt in appointments:
        nombre = folio_map.get(appt["folio"])
        if not nombre:
            unmatched_map[appt["folio"]] = f"folio={appt['folio']} (no en Base de Datos)"
            continue
        athlete_id = name_to_id.get(normalize(nombre))
        if not athlete_id:
            unmatched_map[appt["folio"]] = f"folio={appt['folio']} nombre='{nombre}' (no encontrado en BD)"
            continue
        resolved.append({**appt, "athlete_id": athlete_id, "nombre": nombre})

    unmatched = list(unmatched_map.values())
    print(f"  → {len(resolved)} citas resueltas / {len(unmatched)} folios sin match (x7 meses = {len(unmatched)*7*12} citas omitidas aprox)")
    if unmatched:
        print("  ⚠  Atletas no encontrados en BD:")
        for u in unmatched:
            print(f"     • {u}")

    if not resolved:
        print("No hay citas para insertar. Revisa los nombres de atletas.")
        return

    # 5. Construir registros de eventos
    event_records = []
    for appt in resolved:
        svc = appt["service"] or "SIN SERVICIO"
        start_ts = build_ts(appt["date"], appt["time"])
        end_ts   = build_ts(appt["date"], add_minutes(appt["time"], SLOT_MINUTES))
        event_records.append({
            "title":                svc,
            "event_type":           service_to_event_type(svc),
            "start_at":             start_ts,
            "end_at":               end_ts,
            "status":               "scheduled",
            "created_by_profile_id": admin_profile_id,
            # Auxiliares (se quitan antes de enviar a la API)
            "_folio":               appt["folio"],
            "_athlete_id":          appt["athlete_id"],
        })

    # 6. Insertar eventos en lotes y recoger IDs
    print(f"\nInsertando {len(event_records)} eventos en lotes de {BATCH_SIZE} …")
    inserted_events: list[dict] = []

    for i in range(0, len(event_records), BATCH_SIZE):
        chunk = event_records[i : i + BATCH_SIZE]
        # Quitar campos auxiliares antes de enviar a la API
        api_chunk = [{k: v for k, v in r.items() if not k.startswith("_")} for r in chunk]
        result = sb.insert_batch("events", api_chunk)
        # Re-asociar los IDs devueltos con el athlete_id correspondiente
        for j, row in enumerate(result):
            inserted_events.append({
                "event_id":   row["id"],
                "athlete_id": chunk[j]["_athlete_id"],
            })
        print(f"  [{i + len(chunk)}/{len(event_records)}] eventos insertados", end="\r")
        time.sleep(0.05)  # pequeña pausa para evitar rate limiting

    print(f"\n  ✓ {len(inserted_events)} eventos insertados en la BD")

    # 7. Construir e insertar participantes
    participant_records = [
        {
            "event_id":          ev["event_id"],
            "participant_id":    ev["athlete_id"],
            "participant_type":  "athlete",
            "attendance_status": "planned",
        }
        for ev in inserted_events
    ]

    print(f"Insertando {len(participant_records)} participantes …")
    for i in range(0, len(participant_records), BATCH_SIZE):
        chunk = participant_records[i : i + BATCH_SIZE]
        sb.insert_batch("event_participants", chunk)
        print(f"  [{i + len(chunk)}/{len(participant_records)}] participantes insertados", end="\r")
        time.sleep(0.05)

    print(f"\n  ✓ {len(participant_records)} participantes insertados")

    # 8. Verificar conteo final
    print("\nVerificando conteos en la BD …")
    total_events = sb.count("events", extra_params={"status": "eq.scheduled"})
    total_parts  = sb.count("event_participants")
    print(f"  events             : {total_events}")
    print(f"  event_participants : {total_parts}")

    if unmatched:
        print(f"\n⚠  {len(unmatched)} cita(s) no pudieron insertarse por falta de match.")
        print("   Verifica que los nombres en la BD coincidan con los del Excel.")
    else:
        print("\n✅ Calendario insertado completamente sin errores.")


if __name__ == "__main__":
    main()
