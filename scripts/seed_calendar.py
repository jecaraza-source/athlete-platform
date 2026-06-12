#!/usr/bin/env python3
"""
seed_calendar.py
================
Parsea el Excel del calendario de citas (Junio–Diciembre 2026) y genera un
archivo SQL listo para ejecutar en el SQL Editor de Supabase.

Uso:
    python3 scripts/seed_calendar.py

Salida:
    scripts/seed_calendar.sql   ← pégalo/cópialo en Supabase SQL Editor

Requiere:
    pip install openpyxl  (ya instalado)

Columnas leídas del Excel
    Informacion   → servicio / horario / fecha (DD/MM) / folio de atleta
    Base de Datos → folio → nombre completo del atleta
"""

import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# ──────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────
EXCEL_FILE = (
    "/Users/javierescobedo/Library/Containers/net.whatsapp.WhatsApp"
    "/Data/tmp/documents/84D311DE-F91F-4798-B645-0E33CC12C633"
    "/CALENDARIO CITAS JUNIO-DICIEMBRE_ATLETAS_70626.xlsx"
)
OUTPUT_SQL = Path(__file__).parent / "seed_calendar.sql"

# Duración de cada cita en minutos
SLOT_MINUTES = 30

# Offset de zona horaria (Mexico, UTC-6)
TZ = "-06:00"

# Mapeo servicio → event_type
def service_to_event_type(service: str) -> str:
    s = service.upper()
    if "FISIO" in s:
        return "fisioterapia"
    if "MÉDICO" in s or "MEDICO" in s:
        return "medico"
    if "PSICOL" in s:
        return "psicologia"
    if "NUTRI" in s:
        return "nutricion"
    return "appointment"


# ──────────────────────────────────────────────
# UTILIDADES
# ──────────────────────────────────────────────
def normalize(text: str) -> str:
    """Minúsculas, sin acentos, espacios normalizados."""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    return " ".join(ascii_str.lower().split())


def sql_str(value: str) -> str:
    """Escapa una cadena para SQL (comillas simples)."""
    return value.replace("'", "''")


def build_timestamp(date_str: str, time_str: str) -> str:
    """'2026-06-08' + '13:30' → '2026-06-08T13:30:00-06:00'"""
    return f"{date_str}T{time_str}:00{TZ}"


def add_minutes(time_str: str, minutes: int) -> str:
    """'13:30' + 30 min → '14:00'"""
    h, m = map(int, time_str.split(":"))
    dt = datetime(2000, 1, 1, h, m) + timedelta(minutes=minutes)
    return dt.strftime("%H:%M")


# ──────────────────────────────────────────────
# PARSER DEL EXCEL
# ──────────────────────────────────────────────
def parse_base_de_datos(wb) -> dict:
    """Devuelve {folio_int: nombre_completo_str}"""
    ws = wb["Base de Datos"]
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        folio, nombre = row[0], row[1]
        if folio is None or nombre is None:
            continue
        try:
            folio = int(folio)
        except (ValueError, TypeError):
            continue
        mapping[folio] = str(nombre).strip()
    return mapping


def parse_informacion(wb) -> list:
    """
    Devuelve lista de dicts:
        {service, date (YYYY-MM-DD), time (HH:MM), folio (int)}
    """
    ws = wb["Informacion"]
    appointments = []

    current_year = 2026
    date_cols: dict[int, str] = {}   # col_index → 'DD/MM'
    current_service: str | None = None
    in_data = False
    skip_next = False  # para saltar la fila de nombres de día (L/M/M/J/V)

    for row in ws.iter_rows(values_only=True):
        row = list(row)

        def cell(i):
            v = row[i] if i < len(row) else None
            return str(v).strip() if v is not None else ""

        col0 = cell(0)
        col2 = cell(2)

        # ── Título del mes ──────────────────────────────────────────────
        if "CALENDARIO DE CITAS" in col0:
            m = re.search(r"(\d{4})", col0)
            current_year = int(m.group(1)) if m else current_year
            in_data = False
            date_cols = {}
            current_service = None
            skip_next = False
            continue

        # ── Fila de encabezado SERVICIO ────────────────────────────────
        if col0 == "SERVICIO":
            date_cols = {}
            for ci, val in enumerate(row):
                if val is not None and re.match(r"^\d{2}/\d{2}$", str(val).strip()):
                    date_cols[ci] = str(val).strip()
            in_data = True
            current_service = None
            skip_next = True  # siguiente fila = L/M/M/J/V
            continue

        # ── Saltar fila de nombres de día ──────────────────────────────
        if skip_next:
            skip_next = False
            continue

        # ── Fin de sección mensual ─────────────────────────────────────
        if "TOTAL DE CITAS" in col0:
            in_data = False
            continue

        if not in_data or not date_cols:
            continue

        # ── Actualizar nombre de servicio ─────────────────────────────
        if col0 and col0 not in ("", "None"):
            current_service = col0

        # ── Debe tener un horario válido ──────────────────────────────
        if not re.match(r"^\d{2}:\d{2}$", col2):
            continue

        time_str = col2

        # ── Recorrer columnas de fecha ─────────────────────────────────
        for ci, date_raw in date_cols.items():
            cell_val = row[ci] if ci < len(row) else None
            if cell_val is None:
                continue
            try:
                folio = int(cell_val)
            except (ValueError, TypeError):
                continue

            day, month = date_raw.split("/")
            appointments.append(
                {
                    "service": current_service,
                    "date": f"{current_year}-{month.zfill(2)}-{day.zfill(2)}",
                    "time": time_str,
                    "folio": folio,
                }
            )

    return appointments


# ──────────────────────────────────────────────
# GENERADOR DE SQL
# ──────────────────────────────────────────────
def generate_sql(appointments: list, folio_map: dict) -> str:
    lines = []

    header = """-- ============================================================
-- seed_calendar.sql
-- Calendario de Citas Junio–Diciembre 2026
-- Generado automáticamente — revisar antes de ejecutar.
--
-- INSTRUCCIONES:
--   1. Asegúrate de haber borrado los eventos anteriores si es necesario.
--   2. Pega este archivo en el SQL Editor de Supabase y ejecuta.
--   3. Los RAISE NOTICE al final indican atletas no encontrados.
-- ============================================================

-- Necesaria para comparar nombres con y sin acentos
CREATE EXTENSION IF NOT EXISTS unaccent;

DO $$
DECLARE
  v_event_id   uuid;
  v_athlete_id uuid;
  v_unmatched  int := 0;
BEGIN
"""
    lines.append(header)

    unmatched_folios: set[int] = set()

    for appt in appointments:
        folio   = appt["folio"]
        service = appt["service"] or "SIN SERVICIO"
        date    = appt["date"]
        time    = appt["time"]
        nombre  = folio_map.get(folio)

        start_ts = build_timestamp(date, time)
        end_ts   = build_timestamp(date, add_minutes(time, SLOT_MINUTES))
        ev_type  = service_to_event_type(service)
        title    = sql_str(service)

        if nombre is None:
            unmatched_folios.add(folio)
            lines.append(
                f"  -- FOLIO {folio}: no está en Base de Datos → OMITIDO\n"
            )
            continue

        nombre_sql    = sql_str(nombre)
        nombre_norm   = sql_str(normalize(nombre))

        lines.append(
            f"  -- {service} | {date} {time} | folio={folio} | {nombre}\n"
            f"  SELECT id INTO v_athlete_id\n"
            f"    FROM public.athletes\n"
            f"   WHERE lower(unaccent(trim(first_name || ' ' || last_name)))\n"
            f"       = lower(unaccent('{nombre_norm}'))\n"
            f"   LIMIT 1;\n"
            f"\n"
            f"  IF v_athlete_id IS NULL THEN\n"
            f"    v_unmatched := v_unmatched + 1;\n"
            f"    RAISE NOTICE 'Atleta no encontrado: folio=% nombre=%', {folio}, '{nombre_sql}';\n"
            f"  ELSE\n"
            f"    INSERT INTO public.events (title, event_type, start_at, end_at, status)\n"
            f"    VALUES ('{title}', '{ev_type}', '{start_ts}', '{end_ts}', 'scheduled')\n"
            f"    RETURNING id INTO v_event_id;\n"
            f"\n"
            f"    INSERT INTO public.event_participants\n"
            f"           (event_id, participant_id, participant_type, attendance_status)\n"
            f"    VALUES (v_event_id, v_athlete_id, 'athlete', 'planned');\n"
            f"  END IF;\n"
            f"\n"
        )

    footer = (
        f"  RAISE NOTICE '=== FIN: % citas NO encontradas ===',"
        f" v_unmatched;\n"
        f"END $$;\n"
    )
    lines.append(footer)

    return "".join(lines)


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
def main():
    print(f"Leyendo: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE)

    print("Parseando Base de Datos …")
    folio_map = parse_base_de_datos(wb)
    print(f"  → {len(folio_map)} atletas en Base de Datos")

    print("Parseando Informacion …")
    appointments = parse_informacion(wb)
    print(f"  → {len(appointments)} citas encontradas")

    # Stats
    missing = {a["folio"] for a in appointments if a["folio"] not in folio_map}
    if missing:
        print(f"  ⚠  Folios sin nombre en Base de Datos: {sorted(missing)}")

    print("Generando SQL …")
    sql = generate_sql(appointments, folio_map)

    OUTPUT_SQL.write_text(sql, encoding="utf-8")
    print(f"  → Escrito en: {OUTPUT_SQL}")
    print()
    print("Próximo paso:")
    print("  1. Abre Supabase → SQL Editor")
    print("  2. Copia/pega el contenido de scripts/seed_calendar.sql")
    print("  3. Ejecuta y revisa los RAISE NOTICE al final")


if __name__ == "__main__":
    main()
