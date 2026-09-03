#!/usr/bin/env python3
"""Import morning appointments from the supplied workbook without touching afternoons.

Usage:
  python3 scripts/import_morning_appointments.py --dry-run
  python3 scripts/import_morning_appointments.py --execute
"""

from __future__ import annotations

import argparse
import json
import unicodedata
import uuid
from collections import Counter
from datetime import datetime, time, timedelta, timezone
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path("/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/CALENDARIO ATENCION MATUTINO.xlsx")
ENV_FILE = ROOT / ".env.local"
IMPORT_MARKER = "[MORNING_APPOINTMENT_IMPORT]"
MEXICO_TZ = timezone(timedelta(hours=-6))
BATCH_SIZE = 200
EVENT_TYPES = {"MÉDICO": "medical", "NUTRICIÓN": "nutrition", "PSICOLOGÍA": "psychology"}


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if "=" in line and not line.lstrip().startswith("#"):
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def normalize(value: str) -> str:
    return " ".join(unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower().split())


class Supabase:
    def __init__(self, url: str, key: str) -> None:
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def fetch_all(self, table: str, params: dict[str, str]) -> list[dict]:
        rows: list[dict] = []
        offset = 0
        while True:
            response = requests.get(f"{self.url}/rest/v1/{table}", headers=self.headers, params={**params, "limit": 1000, "offset": offset}, timeout=60)
            response.raise_for_status()
            page = response.json()
            rows.extend(page)
            if len(page) < 1000:
                return rows
            offset += 1000

    def insert(self, table: str, rows: list[dict]) -> list[dict]:
        response = requests.post(
            f"{self.url}/rest/v1/{table}",
            headers={**self.headers, "Prefer": "return=representation"},
            data=json.dumps(rows, ensure_ascii=False).encode(),
            timeout=60,
        )
        response.raise_for_status()
        return response.json()


def parse_workbook() -> list[dict]:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    folio_names = {
        int(folio): str(name).strip()
        for folio, name, *_ in workbook["Base de Datos"].iter_rows(values_only=True)
        if str(folio).strip().isdigit() and name
    }
    sheet = workbook["Folios-Horarios-Servicios"]
    appointments: list[dict] = []
    for title_row in range(1, sheet.max_row + 1):
        title = sheet.cell(title_row, 1).value
        if not isinstance(title, str) or not title.startswith("CALENDARIO DE CITAS"):
            continue
        year = int(title.rsplit(" ", 1)[1])
        for row in range(title_row + 4, min(title_row + 27, sheet.max_row + 1)):
            service, professional, slot_time = (sheet.cell(row, column).value for column in (1, 2, 3))
            if not (service and professional and isinstance(slot_time, time)):
                continue
            service = str(service).strip()
            if service not in EVENT_TYPES:
                raise ValueError(f"Servicio no soportado: {service!r}")
            for column in range(5, sheet.max_column + 1):
                folio, date_label = sheet.cell(row, column).value, sheet.cell(title_row + 2, column).value
                if not isinstance(folio, (int, float)) or not date_label:
                    continue
                folio = int(folio)
                if folio not in folio_names:
                    raise ValueError(f"Folio {folio} no existe en Base de Datos")
                day, month = map(int, str(date_label).split("/"))
                appointments.append({
                    "folio": folio,
                    "name": folio_names[folio],
                    "service": service,
                    "professional": str(professional).strip(),
                    "start_at": datetime(year, month, day, slot_time.hour, slot_time.minute, tzinfo=MEXICO_TZ),
                })
    if len(appointments) != 228:
        raise ValueError(f"Se esperaban 228 citas; se encontraron {len(appointments)}")
    return appointments


def insert_batches(client: Supabase, table: str, rows: list[dict]) -> list[dict]:
    result: list[dict] = []
    for offset in range(0, len(rows), BATCH_SIZE):
        result.extend(client.insert(table, rows[offset : offset + BATCH_SIZE]))
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--execute", action="store_true")
    args = parser.parse_args()
    if args.dry_run == args.execute:
        parser.error("Usa exactamente una opción: --dry-run o --execute")

    appointments = parse_workbook()
    env = load_env(ENV_FILE)
    client = Supabase(env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"])
    athlete_ids = {
        normalize(f"{athlete['first_name']} {athlete['last_name']}"): athlete["id"]
        for athlete in client.fetch_all("athletes", {"select": "id,first_name,last_name"})
    }
    existing = client.fetch_all("events", {
        "select": "description",
        "shift": "eq.morning",
        "description": f"ilike.{IMPORT_MARKER}%",
    })
    existing_sources = {
        row["description"].split("source=", 1)[1].split(";", 1)[0]
        for row in existing
        if row.get("description") and "source=" in row["description"]
    }
    profiles = client.fetch_all("profiles", {"select": "id", "order": "created_at"})
    if not profiles:
        raise ValueError("No hay perfil disponible para registrar la importación")

    events: list[dict] = []
    athlete_ids_for_events: list[str] = []
    unresolved: list[str] = []
    for appointment in appointments:
        athlete_id = athlete_ids.get(normalize(appointment["name"]))
        source = "|".join((str(appointment["folio"]), appointment["service"], appointment["professional"], appointment["start_at"].isoformat()))
        if not athlete_id:
            unresolved.append(f"{appointment['folio']}: {appointment['name']}")
        elif source not in existing_sources:
            start_at = appointment["start_at"]
            events.append({
                "id": str(uuid.uuid4()),
                "title": appointment["professional"],
                "event_type": EVENT_TYPES[appointment["service"]],
                "start_at": start_at.isoformat(),
                "end_at": (start_at + timedelta(minutes=20)).isoformat(),
                "status": "scheduled",
                "shift": "morning",
                "created_by_profile_id": profiles[0]["id"],
                "description": f"{IMPORT_MARKER} source={source}; Folio: {appointment['folio']}; Turno: morning",
            })
            athlete_ids_for_events.append(athlete_id)

    if unresolved:
        raise ValueError(f"Atletas sin resolver: {', '.join(sorted(unresolved))}")
    print(f"Citas en Excel: {len(appointments)}")
    print(f"Ya importadas (omitidas): {len(appointments) - len(events)}")
    print(f"Eventos matutinos a insertar: {len(events)}")
    print(f"Servicios: {dict(Counter(item['service'] for item in appointments))}")
    if args.dry_run:
        print("Validación completa; no se escribió ningún dato.")
        return

    inserted = insert_batches(client, "events", events)
    participants = [{
        "event_id": event["id"],
        "participant_id": athlete_id,
        "participant_type": "athlete",
        "attendance_status": "planned",
    } for event, athlete_id in zip(inserted, athlete_ids_for_events)]
    insert_batches(client, "event_participants", participants)
    print(f"Importación completada: {len(inserted)} eventos matutinos y participantes.")


if __name__ == "__main__":
    main()
