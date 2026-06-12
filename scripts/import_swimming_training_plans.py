#!/usr/bin/env python3
"""
import_swimming_training_plans.py
----------------------------------
Reads each sheet from the individualized swimming training plan Excel file and
upserts the data into the `athlete_coach_evaluation` table in Supabase.

Mapping logic:
  - Matches each sheet's athlete against the `athletes` table by normalized full name
  - Creates `athlete_initial_diagnostic` + `athlete_diagnostic_sections` (section='entrenador')
    if they don't already exist
  - Upserts `athlete_coach_evaluation` with all available fields

Run from the repo root:
    python3 scripts/import_swimming_training_plans.py
"""

import unicodedata
import sys
import os
import json
import requests
import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = (
    "/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/"
    "Entrenamientos/Formato_Plan_Entrenamiento_V1.0 Natacion 2.xlsx"
)

SUPABASE_URL = "https://gwjnqokwchdojlcngtbi.supabase.co"
SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SERVICE_ROLE_KEY:
    # Fall back to reading the .env.local for convenience when run locally
    env_path = os.path.join(os.path.dirname(__file__), "../apps/web/.env.local")
    try:
        with open(env_path) as f:
            for line in f:
                if line.startswith("SUPABASE_SERVICE_ROLE_KEY="):
                    SERVICE_ROLE_KEY = line.strip().split("=", 1)[1]
    except FileNotFoundError:
        pass

if not SERVICE_ROLE_KEY:
    print("ERROR: SUPABASE_SERVICE_ROLE_KEY not found.", file=sys.stderr)
    sys.exit(1)

HEADERS = {
    "apikey": SERVICE_ROLE_KEY,
    "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

BASE = f"{SUPABASE_URL}/rest/v1"

# All 5 section keys that must exist per diagnostic
ALL_SECTION_KEYS = ["medico", "nutricion", "psicologia", "entrenador", "fisioterapia"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize(text: str) -> str:
    """Lower-case, strip accents, collapse whitespace."""
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_str = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(ascii_str.lower().split())


def rq(method: str, path: str, **kwargs):
    """Thin wrapper around requests with error propagation."""
    resp = getattr(requests, method)(f"{BASE}/{path}", headers=HEADERS, **kwargs)
    if not resp.ok:
        raise RuntimeError(f"{method.upper()} /{path} → {resp.status_code}: {resp.text[:400]}")
    # 204 No Content
    if resp.status_code == 204:
        return None
    return resp.json()


def get_all_athletes():
    """Return list of {id, first_name, last_name, norm_name}."""
    data = rq("get", "athletes", params={"select": "id,first_name,last_name"})
    athletes = []
    for a in (data or []):
        full = f"{a['first_name']} {a['last_name']}"
        athletes.append({**a, "norm_name": normalize(full)})
    return athletes


def find_athlete(athletes, full_name_from_excel: str):
    """
    Try exact norm match first, then look for any athlete whose norm_name
    tokens are all present in the Excel full name (or vice-versa).
    """
    target = normalize(full_name_from_excel)
    # 1. Exact match
    for a in athletes:
        if a["norm_name"] == target:
            return a
    # 2. Partial match: all DB tokens are in the target
    for a in athletes:
        db_tokens = set(a["norm_name"].split())
        if db_tokens and db_tokens.issubset(set(target.split())):
            return a
    # 3. Partial match: all target tokens are in DB name
    for a in athletes:
        target_tokens = set(target.split())
        if target_tokens and target_tokens.issubset(set(a["norm_name"].split())):
            return a
    return None


def ensure_diagnostic(athlete_id: str) -> str:
    """
    Get or create the athlete_initial_diagnostic row.
    Returns the diagnostic id.
    """
    existing = rq("get", "athlete_initial_diagnostic",
                  params={"athlete_id": f"eq.{athlete_id}", "select": "id"})
    if existing:
        return existing[0]["id"]

    # Create diagnostic header
    created = rq("post", "athlete_initial_diagnostic",
                 json={"athlete_id": athlete_id})
    diag_id = created[0]["id"]

    # Seed all 5 section rows
    sections_payload = [
        {"diagnostic_id": diag_id, "athlete_id": athlete_id, "section": s}
        for s in ALL_SECTION_KEYS
    ]
    rq("post", "athlete_diagnostic_sections", json=sections_payload)
    return diag_id


def ensure_section(athlete_id: str, diag_id: str, section: str) -> str:
    """
    Get or create the athlete_diagnostic_sections row for 'entrenador'.
    Returns the section row id.
    """
    existing = rq("get", "athlete_diagnostic_sections",
                  params={
                      "athlete_id": f"eq.{athlete_id}",
                      "section": f"eq.{section}",
                      "select": "id",
                  })
    if existing:
        return existing[0]["id"]

    created = rq("post", "athlete_diagnostic_sections",
                 json={"diagnostic_id": diag_id, "athlete_id": athlete_id, "section": section})
    return created[0]["id"]


def mark_section_complete(section_id: str, diag_id: str):
    """Update section status to completo and recalc overall."""
    import datetime
    now = datetime.datetime.utcnow().isoformat() + "Z"
    rq("patch", "athlete_diagnostic_sections",
       params={"id": f"eq.{section_id}"},
       json={"status": "completo", "completion_pct": 100, "completed_at": now,
             "captured_at": now, "updated_at": now})

    # Recalculate overall status
    sections = rq("get", "athlete_diagnostic_sections",
                  params={"diagnostic_id": f"eq.{diag_id}",
                          "select": "status,completion_pct"})
    if not sections:
        return
    total = len(sections)
    completed = sum(1 for s in sections if s["status"] == "completo")
    has_attn  = any(s["status"] == "requiere_atencion" for s in sections)
    in_proc   = any(s["status"] == "en_proceso" for s in sections)
    pct       = round(completed / total * 100) if total else 0

    if has_attn:
        overall = "requiere_atencion"
    elif completed == total and total > 0:
        overall = "completo"
    elif completed > 0 or in_proc:
        overall = "en_proceso"
    else:
        overall = "pendiente"

    patch = {"completion_pct": pct, "overall_status": overall, "updated_at": now}
    if overall == "completo":
        patch["completed_at"] = now
    rq("patch", "athlete_initial_diagnostic",
       params={"id": f"eq.{diag_id}"}, json=patch)


def upsert_coach_evaluation(section_id: str, athlete_id: str, payload: dict):
    """
    Insert or update the athlete_coach_evaluation row.
    Fetches existing row first; PATCHes if found, POSTs if not.
    """
    import datetime
    now = datetime.datetime.utcnow().isoformat() + "Z"

    # Check if a row already exists for this section
    existing = rq("get", "athlete_coach_evaluation",
                  params={"diagnostic_section_id": f"eq.{section_id}",
                          "select": "id"})

    data = {"updated_at": now, **payload}

    if existing:
        # UPDATE existing row
        resp = requests.patch(
            f"{BASE}/athlete_coach_evaluation",
            headers={**HEADERS, "Prefer": "return=representation"},
            params={"diagnostic_section_id": f"eq.{section_id}"},
            json=data,
        )
    else:
        # INSERT new row
        data["diagnostic_section_id"] = section_id
        data["athlete_id"] = athlete_id
        resp = requests.post(
            f"{BASE}/athlete_coach_evaluation",
            headers={**HEADERS, "Prefer": "return=representation"},
            json=data,
        )

    if not resp.ok:
        raise RuntimeError(f"upsert coach eval → {resp.status_code}: {resp.text[:400]}")


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_sheet(ws) -> dict:
    """
    Extract all label→value pairs from a sheet.
    Returns a dict with keys matching the label text in column A, values from column C.
    """
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        label = row[0]
        value = row[2]  # column C
        if label and isinstance(label, str) and label.strip() and value is not None:
            key = label.strip()
            data[key] = str(value).strip() if value is not None else None
    return data


def map_to_coach_eval(raw: dict) -> dict:
    """Map raw Excel label→value to athlete_coach_evaluation column names."""
    def v(key): return raw.get(key) or None

    # For discipline_intervention: combine generic blocks + specific plan
    blocks = v("Bloques de intervención por disciplina")
    specific = v("Plan de intervención específico para la disciplina del atleta")
    if blocks and specific:
        discipline_intervention = f"{blocks}\n\n{specific}"
    else:
        discipline_intervention = specific or blocks

    return {
        # 1. Evaluación Inicial — Pruebas Físicas
        "strength_test":            v("Fuerza"),
        "power_test":               v("Potencia"),
        "speed_test":               v("Velocidad"),
        "endurance_test":           v("Resistencia"),
        "flexibility_test":         v("Flexibilidad"),
        # 2. Análisis Técnico
        "technical_weaknesses":     v("Detección de debilidades técnicas"),
        "competitive_capabilities": v("Valoración de capacidades competitivas"),
        # 3. Evaluación Biomecánica
        "movement_efficiency":      v("Eficiencia del movimiento"),
        "body_mechanics":           v("Mecánica corporal"),
        "segment_alignment":        v("Alineación segmentaria"),
        # 4. Perfil Deportivo
        "athlete_sport_profile":    v("Perfil deportivo completo"),
        # 5. Intervención por Disciplina
        "discipline_intervention":  discipline_intervention,
        # 6. Plan de Entrenamiento Individualizado
        "season_structure":         v("Estructura de la temporada deportiva"),
        "competitive_calendar":     v("Calendario competitivo"),
        "performance_objectives":   v("Objetivos de rendimiento"),
        "preparation_stages":       v("Etapas de preparación del atleta"),
        # 7. Supervisión del Entrenador
        "technical_correction":     v("Corrección técnica"),
        "load_supervision":         v("Supervisión de cargas de entrenamiento"),
        "competition_preparation":  v("Preparación para competencia"),
        "performance_analysis":     v("Análisis de desempeño"),
        "continuous_feedback":      v("Retroalimentación continua"),
        "mark_monitoring":          v("Monitoreo de mejora en marcas"),
        "plan_adjustments":         v("Ajuste del plan de entrenamiento"),
        "observations":             v("Observaciones Generales"),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print("Fetching athletes from Supabase…")
    athletes = get_all_athletes()
    print(f"  Found {len(athletes)} athletes in DB")

    results = {"matched": [], "unmatched": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw = parse_sheet(ws)

        full_name = raw.get("Nombre Completo del Atleta", "").strip()
        if not full_name:
            print(f"  [{sheet_name}] No athlete name found — skipping")
            continue

        athlete = find_athlete(athletes, full_name)
        if not athlete:
            print(f"  [{sheet_name}] ⚠️  NO MATCH for '{full_name}'")
            results["unmatched"].append({"sheet": sheet_name, "name": full_name})
            continue

        print(f"  [{sheet_name}] Matched → {athlete['first_name']} {athlete['last_name']} (id={athlete['id'][:8]}…)")

        try:
            # 1. Ensure diagnostic header
            diag_id = ensure_diagnostic(athlete["id"])

            # 2. Ensure entrenador section
            section_id = ensure_section(athlete["id"], diag_id, "entrenador")

            # 3. Build & upsert coach evaluation
            payload = map_to_coach_eval(raw)
            upsert_coach_evaluation(section_id, athlete["id"], payload)

            # 4. Mark section complete and recalc overall
            mark_section_complete(section_id, diag_id)

            results["matched"].append({
                "sheet": sheet_name,
                "name": full_name,
                "athlete_id": athlete["id"],
            })
            print(f"    ✓ Coach evaluation upserted")

        except Exception as e:
            print(f"    ✗ ERROR: {e}")
            results["unmatched"].append({"sheet": sheet_name, "name": full_name, "error": str(e)})

    print("\n" + "=" * 60)
    print(f"Done. Matched: {len(results['matched'])}  |  Unmatched/Error: {len(results['unmatched'])}")
    if results["unmatched"]:
        print("\nUnmatched or errored:")
        for u in results["unmatched"]:
            print(f"  • {u}")


if __name__ == "__main__":
    main()
