"""
Upload TKD training plans from Excel files to Supabase.

Maps each sheet → athlete (by name matching), then upserts into:
  - athlete_initial_diagnostic      (created if missing)
  - athlete_diagnostic_sections     (all 5 sections seeded if missing)
  - athlete_coach_evaluation        (upserted on diagnostic_section_id)

Sets section status to 'en_proceso' and recalculates the overall diagnostic status.
"""

import os
import re
import sys
import unicodedata
from datetime import datetime, timezone

import openpyxl
from supabase import create_client

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]

FILES = [
    "/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE TKD Dos.xlsx",
    "/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/Entrenamientos/PE TKD Uno.xlsx",
]

# Excel row number (1-indexed) → coach-evaluation field name
# Column C (index 3) holds the value in every case.
ROW_TO_FIELD = {
    10: "strength_test",
    13: "power_test",
    16: "speed_test",
    19: "endurance_test",
    22: "flexibility_test",
    26: "technical_weaknesses",
    29: "competitive_capabilities",
    33: "movement_efficiency",
    36: "body_mechanics",
    39: "segment_alignment",
    43: "athlete_sport_profile",
    50: "discipline_intervention",   # row 47 is static UI text, skip it
    54: "season_structure",
    57: "competitive_calendar",
    60: "performance_objectives",
    63: "preparation_stages",
    67: "technical_correction",
    70: "load_supervision",
    73: "competition_preparation",
    76: "performance_analysis",
    79: "continuous_feedback",
    82: "mark_monitoring",
    85: "plan_adjustments",
    88: "observations",
}

SECTION_KEYS = ["medico", "nutricion", "psicologia", "entrenador", "fisioterapia"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize(text: str) -> str:
    """Lowercase, strip accents and extra whitespace."""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text)


def fuzzy_match(db_athletes: list[dict], full_name: str) -> dict | None:
    """Return the best-matching DB athlete by word-overlap score (threshold 0.4)."""
    needle_words = set(normalize(full_name).split())
    best, best_score = None, 0.0
    for a in db_athletes:
        haystack = normalize(f"{a['first_name']} {a['last_name']}")
        haystack_words = set(haystack.split())
        common = needle_words & haystack_words
        score = len(common) / max(len(needle_words), len(haystack_words))
        if score > best_score:
            best_score = score
            best = a
    return best if best_score >= 0.4 else None


def parse_sheet(ws) -> dict:
    """Extract athlete metadata and all training-plan fields from a worksheet."""
    def cell_str(row: int, col: int = 3) -> str | None:
        val = ws.cell(row=row, column=col).value
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    fields = {}
    for row_idx, field_name in ROW_TO_FIELD.items():
        fields[field_name] = cell_str(row_idx)

    return {
        "full_name": cell_str(3),
        "rama":      cell_str(5),   # 'femenil' / 'varonil'
        "fields":    fields,
    }


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Supabase operations
# ---------------------------------------------------------------------------

def ensure_diagnostic_and_section(sb, athlete_id: str) -> dict:
    """
    Guarantee:
      1. One row in athlete_initial_diagnostic for the athlete.
      2. All 5 section rows in athlete_diagnostic_sections.
    Returns the 'entrenador' section row  {id, diagnostic_id}.
    """
    # 1. Diagnostic header
    res = sb.table("athlete_initial_diagnostic").select("id").eq("athlete_id", athlete_id).execute()
    if res.data:
        diag_id = res.data[0]["id"]
    else:
        ins = sb.table("athlete_initial_diagnostic").insert({"athlete_id": athlete_id}).select("id").execute()
        diag_id = ins.data[0]["id"]
        # Seed all 5 sections
        sb.table("athlete_diagnostic_sections").insert([
            {"diagnostic_id": diag_id, "athlete_id": athlete_id, "section": s}
            for s in SECTION_KEYS
        ]).execute()

    # 2. Coach section
    sec_res = (
        sb.table("athlete_diagnostic_sections")
        .select("id, diagnostic_id")
        .eq("athlete_id", athlete_id)
        .eq("section", "entrenador")
        .execute()
    )
    if sec_res.data:
        return sec_res.data[0]

    # Section missing (older diagnostic) – create it individually
    new_sec = (
        sb.table("athlete_diagnostic_sections")
        .insert({"diagnostic_id": diag_id, "athlete_id": athlete_id, "section": "entrenador"})
        .select("id, diagnostic_id")
        .execute()
    )
    return new_sec.data[0]


def upload_coach_evaluation(sb, athlete_id: str, section_id: str, fields: dict) -> None:
    payload = {
        "diagnostic_section_id": section_id,
        "athlete_id": athlete_id,
        **fields,
        "updated_at": now_iso(),
    }
    sb.table("athlete_coach_evaluation").upsert(
        payload, on_conflict="diagnostic_section_id"
    ).execute()


def update_section_and_recalculate(sb, section_id: str, diagnostic_id: str, athlete_id: str) -> None:
    now = now_iso()
    sb.table("athlete_diagnostic_sections").update({
        "status":         "en_proceso",
        "completion_pct": 50,
        "updated_at":     now,
        "captured_at":    now,
    }).eq("id", section_id).execute()

    # Recalculate overall diagnostic
    secs = (
        sb.table("athlete_diagnostic_sections")
        .select("status")
        .eq("diagnostic_id", diagnostic_id)
        .execute()
    )
    if not secs.data:
        return

    total     = len(secs.data)
    completed = sum(1 for s in secs.data if s["status"] == "completo")
    has_attn  = any(s["status"] == "requiere_atencion" for s in secs.data)
    in_proc   = any(s["status"] == "en_proceso" for s in secs.data)
    pct       = round((completed / total) * 100) if total else 0

    if has_attn:
        overall = "requiere_atencion"
    elif completed == total and total > 0:
        overall = "completo"
    elif completed > 0 or in_proc:
        overall = "en_proceso"
    else:
        overall = "pendiente"

    sb.table("athlete_initial_diagnostic").update({
        "completion_pct": pct,
        "overall_status": overall,
        "updated_at":     now,
    }).eq("id", diagnostic_id).execute()


def create_athlete(sb, full_name: str, rama: str | None) -> dict:
    parts = full_name.title().split()
    first_name = " ".join(parts[:2]) if len(parts) >= 2 else parts[0]
    last_name  = " ".join(parts[2:]) if len(parts) > 2 else ""
    sex = None
    if rama:
        if "varonil" in rama.lower():
            sex = "male"
        elif "femenil" in rama.lower():
            sex = "female"
    ins = sb.table("athletes").insert({
        "first_name": first_name,
        "last_name":  last_name,
        "discipline": "taekwondo",
        "status":     "active",
        **({"sex": sex} if sex else {}),
    }).select("id, first_name, last_name").execute()
    return ins.data[0]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Load all taekwondo athletes from DB once
    db_athletes = (
        sb.table("athletes")
        .select("id, first_name, last_name")
        .eq("discipline", "taekwondo")
        .execute()
        .data
    )
    print(f"Loaded {len(db_athletes)} taekwondo athletes from DB.\n")

    successes, failures = [], []

    for file_path in FILES:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"─── {file_path.split('/')[-1]} ───")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 10:       # skip blank/empty sheets
                continue

            parsed = parse_sheet(ws)
            full_name = parsed["full_name"]
            if not full_name:
                continue

            print(f"\n  Sheet: {sheet_name}  →  Excel name: \"{full_name}\"")

            # --- Match or create athlete ---
            athlete = fuzzy_match(db_athletes, full_name)

            if athlete:
                print(f"    ✔ Matched DB: {athlete['first_name']} {athlete['last_name']}  [{athlete['id']}]")
            else:
                print(f"    ✦ No match found — creating new athlete…")
                athlete = create_athlete(sb, full_name, parsed["rama"])
                db_athletes.append(athlete)   # update local cache
                print(f"    ✔ Created: {athlete['first_name']} {athlete['last_name']}  [{athlete['id']}]")

            try:
                section = ensure_diagnostic_and_section(sb, athlete["id"])
                upload_coach_evaluation(sb, athlete["id"], section["id"], parsed["fields"])
                update_section_and_recalculate(sb, section["id"], section["diagnostic_id"], athlete["id"])
                print(f"    ✔ Coach evaluation saved (section {section['id'][:8]}…)")
                successes.append(f"{sheet_name} → {athlete['first_name']} {athlete['last_name']}")
            except Exception as exc:
                msg = f"{sheet_name} → {full_name}: {exc}"
                print(f"    ✘ ERROR: {exc}", file=sys.stderr)
                failures.append(msg)

    print(f"\n{'='*60}")
    print(f"Done.  ✔ {len(successes)} succeeded  ✘ {len(failures)} failed")
    for s in successes:
        print(f"  ✔ {s}")
    if failures:
        print()
        for f in failures:
            print(f"  ✘ {f}", file=sys.stderr)


if __name__ == "__main__":
    main()
