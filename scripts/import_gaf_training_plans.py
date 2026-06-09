#!/usr/bin/env python3
"""
import_gaf_training_plans.py
-----------------------------
Reads PE GAF.xlsx (Gimnasia Artística Femenil training plans) and imports each
athlete's individualised plan into the Supabase `plans` + `athlete_plans` tables.

Usage:
    python3 scripts/import_gaf_training_plans.py

Requirements:
    pip install openpyxl supabase

Env vars (loaded from apps/web/.env.local):
    NEXT_PUBLIC_SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY
"""

import os
import re
import sys
import json
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path
from datetime import datetime, timezone

import openpyxl
from supabase import create_client, Client

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

EXCEL_PATH = Path(
    "/Users/javierescobedo/Library/CloudStorage/Dropbox/"
    "AO Deporte/Entrenamientos/PE GAF.xlsx"
)

# Load env vars from .env.local without requiring python-dotenv
ENV_FILE = Path(__file__).parent.parent / "apps" / "web" / ".env.local"


def load_env(path: Path) -> dict:
    env: dict = {}
    if not path.exists():
        return env
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        # Strip surrounding quotes if present
        val = val.strip().strip('"').strip("'")
        env[key.strip()] = val
    return env


env = load_env(ENV_FILE)

SUPABASE_URL = env.get("NEXT_PUBLIC_SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
SERVICE_KEY  = env.get("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SERVICE_KEY:
    sys.exit(
        "ERROR: NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY "
        "must be set (checked .env.local and environment)."
    )

# ---------------------------------------------------------------------------
# Excel parsing helpers
# ---------------------------------------------------------------------------

def val(ws, key: str) -> str | None:
    """Return the value from column C for the row whose column A matches key."""
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == key:
            raw = row[2]  # column C (0-indexed: 2)
            if raw is None:
                return None
            return str(raw).strip() if not isinstance(raw, str) else raw.strip()
    return None


def parse_sheet(ws) -> dict:
    """Parse a single athlete worksheet into a structured dict."""
    return {
        "categoria":    val(ws, "Categoría / Prueba"),
        "edad":         val(ws, "Edad"),
        "evaluacion_fisica": {
            "fuerza":       val(ws, "Fuerza"),
            "potencia":     val(ws, "Potencia"),
            "velocidad":    val(ws, "Velocidad"),
            "resistencia":  val(ws, "Resistencia"),
            "flexibilidad": val(ws, "Flexibilidad"),
        },
        "analisis_tecnico": {
            "debilidades":               val(ws, "Detección de debilidades técnicas"),
            "capacidades_competitivas":  val(ws, "Valoración de capacidades competitivas"),
        },
        "evaluacion_biomecanica": {
            "eficiencia": val(ws, "Eficiencia del movimiento"),
            "mecanica":   val(ws, "Mecánica corporal"),
            "alineacion": val(ws, "Alineación segmentaria"),
        },
        "perfil_deportivo": val(ws, "Perfil deportivo completo"),
        "intervencion": {
            "bloques":         val(ws, "Bloques de intervención por disciplina"),
            "plan_especifico": val(ws, "Plan de intervención específico para la disciplina del atleta"),
        },
        "plan_individualizado": {
            "estructura_temporada":    val(ws, "Estructura de la temporada deportiva"),
            "calendario_competitivo":  val(ws, "Calendario competitivo"),
            "objetivos_rendimiento":   val(ws, "Objetivos de rendimiento"),
            "etapas_preparacion":      val(ws, "Etapas de preparación del atleta"),
        },
        "supervision_entrenador": {
            "correccion_tecnica":      val(ws, "Corrección técnica"),
            "supervision_cargas":      val(ws, "Supervisión de cargas de entrenamiento"),
            "preparacion_competencia": val(ws, "Preparación para competencia"),
            "analisis_desempeno":      val(ws, "Análisis de desempeño"),
            "retroalimentacion":       val(ws, "Retroalimentación continua"),
            "monitoreo_marcas":        val(ws, "Monitoreo de mejora en marcas"),
            "ajuste_plan":             val(ws, "Ajuste del plan de entrenamiento"),
        },
        "observaciones_generales": val(ws, "Observaciones Generales"),
    }


def normalize(name: str) -> str:
    """Uppercase, collapse whitespace, and strip accents for comparison."""
    # Strip accents: decompose to NFD then keep only ASCII chars
    nfd = unicodedata.normalize("NFD", name)
    ascii_only = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    result = ascii_only.upper().strip()
    return re.sub(r"\s+", " ", result)


# ---------------------------------------------------------------------------
# Name-matching helper
# ---------------------------------------------------------------------------

def find_athlete(athletes: list[dict], full_name: str) -> dict | None:
    """
    Match an athlete record by normalised full name.
    Strategies (in order):
      1. Exact match after accent-stripping + whitespace collapse
      2. Prefix match (Excel sheet names truncated at 31 chars)
      3. Token overlap: all tokens of the shorter name appear in the longer
    """
    needle = normalize(full_name)
    needle_tokens = set(needle.split())

    # Pass 1: exact
    for a in athletes:
        if normalize(f"{a['first_name']} {a['last_name']}") == needle:
            return a

    # Pass 2: prefix (handles 31-char Excel sheet name truncation)
    for a in athletes:
        db_name = normalize(f"{a['first_name']} {a['last_name']}")
        if db_name.startswith(needle) or needle.startswith(db_name):
            return a

    # Pass 3: token overlap — every token in the shorter name appears in the longer
    for a in athletes:
        db_name = normalize(f"{a['first_name']} {a['last_name']}")
        db_tokens = set(db_name.split())
        shorter, longer = (
            (needle_tokens, db_tokens)
            if len(needle_tokens) <= len(db_tokens)
            else (db_tokens, needle_tokens)
        )
        if shorter and shorter.issubset(longer):
            return a

    # Pass 4: fuzzy full-name similarity (handles single-letter typos like Isabella/Isabela)
    best_score, best_match = 0.0, None
    for a in athletes:
        db_name = normalize(f"{a['first_name']} {a['last_name']}")
        score = SequenceMatcher(None, needle, db_name).ratio()
        if score > best_score:
            best_score, best_match = score, a
    if best_score >= 0.90:
        return best_match

    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Connecting to Supabase: {SUPABASE_URL}")
    sb: Client = create_client(SUPABASE_URL, SERVICE_KEY)

    # ── 1. Fetch all active athletes ─────────────────────────────────────────
    res = sb.table("athletes").select("id, first_name, last_name, status").execute()
    athletes = res.data or []
    print(f"Found {len(athletes)} athletes in the database.")

    # ── 2. Load Excel ─────────────────────────────────────────────────────────
    if not EXCEL_PATH.exists():
        sys.exit(f"ERROR: Excel file not found at {EXCEL_PATH}")

    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    print(f"Loaded {len(wb.sheetnames)} sheets from {EXCEL_PATH.name}\n")

    results = {"matched": [], "unmatched": [], "errors": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # The full name is in the sheet itself (row "Nombre Completo del Atleta")
        excel_name = val(ws, "Nombre Completo del Atleta") or sheet_name

        athlete = find_athlete(athletes, excel_name)

        if not athlete:
            print(f"  ⚠️  UNMATCHED: {excel_name!r}")
            results["unmatched"].append(excel_name)
            continue

        plan_data = parse_sheet(ws)
        athlete_id = athlete["id"]
        athlete_db_name = f"{athlete['first_name']} {athlete['last_name']}"

        # ── 3. Upsert plan row ────────────────────────────────────────────────
        #
        # Strategy: insert a new plan only if one with the same title doesn't
        # already exist for this athlete + type combo, so re-runs are idempotent.
        #
        plan_title = f"Plan de Entrenamiento GAF — {excel_name.title()}"

        # Check for existing plan assigned to this athlete with this title
        existing = (
            sb.table("athlete_plans")
            .select("plan_id, plans(id, title)")
            .eq("athlete_id", athlete_id)
            .execute()
        )
        existing_plan_id = None
        for row in (existing.data or []):
            plan_info = row.get("plans")
            if plan_info and plan_info.get("title") == plan_title:
                existing_plan_id = plan_info["id"]
                break

        if existing_plan_id:
            # Update training_plan_data in place
            sb.table("plans").update(
                {
                    "training_plan_data": plan_data,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }
            ).eq("id", existing_plan_id).execute()
            print(f"  ✅  UPDATED  : {athlete_db_name}  →  plan {existing_plan_id[:8]}…")
            results["matched"].append({"athlete": athlete_db_name, "plan_id": existing_plan_id, "action": "updated"})
            continue

        # Insert new plan
        now = datetime.now(timezone.utc).isoformat()
        insert_res = (
            sb.table("plans")
            .insert(
                {
                    "type":               "training",
                    "title":              plan_title,
                    "description":        (
                        f"Plan de entrenamiento individualizado — "
                        f"Gimnasia Artística Femenil\n"
                        f"Categoría: {plan_data.get('categoria') or 'N/A'}"
                    ),
                    "notes":              plan_data.get("observaciones_generales"),
                    "is_published":       True,
                    "training_plan_data": plan_data,
                    "updated_at":         now,
                }
            )
            .select("id")
            .execute()
        )

        if not insert_res.data:
            msg = f"DB insert failed for {excel_name}"
            print(f"  ❌  ERROR    : {msg}")
            results["errors"].append(msg)
            continue

        plan_id = insert_res.data[0]["id"]

        # ── 4. Link plan → athlete ────────────────────────────────────────────
        sb.table("athlete_plans").upsert(
            {
                "plan_id":         plan_id,
                "athlete_id":      athlete_id,
                "assignment_mode": "individual",
            },
            on_conflict="plan_id, athlete_id",
            ignore_duplicates=True,
        ).execute()

        print(f"  ✅  IMPORTED : {athlete_db_name}  →  plan {plan_id[:8]}…")
        results["matched"].append({"athlete": athlete_db_name, "plan_id": plan_id, "action": "created"})

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"  Matched & imported : {len(results['matched'])}")
    print(f"  Unmatched          : {len(results['unmatched'])}")
    print(f"  Errors             : {len(results['errors'])}")
    if results["unmatched"]:
        print("\nUnmatched athletes (check spelling in DB):")
        for name in results["unmatched"]:
            print(f"  - {name}")
    if results["errors"]:
        print("\nErrors:")
        for err in results["errors"]:
            print(f"  - {err}")
    print("=" * 60)


if __name__ == "__main__":
    main()
