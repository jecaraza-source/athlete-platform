#!/usr/bin/env python3
"""
import_athlete_demographics.py
================================
Importa campos demográficos (CURP, CP, Colonia, Teléfono) desde el Excel
INTEGRADO_ATLETAS_POR_DISCIPLINA.xlsx (hoja CONSOLIDADO) hacia la tabla athletes.

Prerequisito: aplicar primero la migración
  supabase/migrations/20260719000001_athletes_curp_cp_colonia_phone.sql

Uso:
  # Ver qué se actualizaría (sin escribir en BD):
  NEXT_PUBLIC_SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... python3 scripts/import_athlete_demographics.py --dry-run

  # Aplicar cambios:
  NEXT_PUBLIC_SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... python3 scripts/import_athlete_demographics.py

  # Solo mostrar atletas del Excel que no se encontraron en BD:
  ... python3 scripts/import_athlete_demographics.py --dry-run --show-missing

Variables de entorno (también se leen de .env.local si existe):
  NEXT_PUBLIC_SUPABASE_URL  URL del proyecto Supabase
  SUPABASE_SERVICE_ROLE_KEY Service role key (bypass RLS)
"""

import json
import os
import re
import ssl
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

EXCEL_PATH = (
    "/Users/javierescobedo/Library/CloudStorage/Dropbox/AO Deporte/"
    "Entrenamientos/INTEGRADO_ATLETAS_POR_DISCIPLINA.xlsx"
)
SHEET_NAME = "CONSOLIDADO"

# Índices de columna en la hoja CONSOLIDADO (0-based)
COL_FOLIO     = 1   # Folio / athlete_code (AT001, AT002…)
COL_NOMBRE    = 3   # Nombre
COL_APELLIDO  = 4   # Apellido
COL_CURP      = 11  # CURP
COL_CP        = 12  # Código Postal
COL_COLONIA   = 13  # Colonia
COL_TELEFONO  = 14  # Teléfono

DRY_RUN      = "--dry-run"      in sys.argv
SHOW_MISSING = "--show-missing" in sys.argv

SSL_CTX = ssl._create_unverified_context()

# ---------------------------------------------------------------------------
# Cargar credenciales
# ---------------------------------------------------------------------------

def load_env() -> tuple[str, str]:
    env_file = Path(__file__).parent.parent / ".env.local"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, _, v = line.partition("=")
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        sys.exit(
            "❌  Faltan variables de entorno:\n"
            "    NEXT_PUBLIC_SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY\n"
            "    (o defínelas en .env.local)"
        )
    return url.rstrip("/"), key


# ---------------------------------------------------------------------------
# Helpers Supabase REST
# ---------------------------------------------------------------------------

SUPABASE_URL: str = ""
SUPABASE_KEY: str = ""


def _headers(extra: dict | None = None) -> dict:
    h = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }
    if extra:
        h.update(extra)
    return h


def api_get(path: str) -> list:
    req = urllib.request.Request(SUPABASE_URL + path, headers=_headers())
    resp = urllib.request.urlopen(req, context=SSL_CTX)
    return json.loads(resp.read())


def api_patch(path: str, data: dict) -> None:
    body = json.dumps(data).encode()
    req = urllib.request.Request(
        SUPABASE_URL + path,
        data=body,
        headers=_headers({"Prefer": "return=minimal"}),
        method="PATCH",
    )
    try:
        urllib.request.urlopen(req, context=SSL_CTX).read()
    except urllib.error.HTTPError as e:
        print(f"  ⚠️  HTTP {e.code}: {e.read().decode()[:300]}")
        raise


# ---------------------------------------------------------------------------
# Limpieza de valores
# ---------------------------------------------------------------------------

def clean(value) -> str | None:
    """Convierte cualquier valor a string limpio, o None si está vacío/inválido."""
    if value is None:
        return None
    s = str(value).strip()
    # Descartar valores de error de Excel o cadenas vacías
    if s in ("", "#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "None"):
        return None
    return s


def clean_phone(value) -> str | None:
    """Normaliza teléfono: quita espacios internos, conserva + inicial."""
    s = clean(value)
    if not s:
        return None
    # Quitar espacios y guiones internos pero conservar el + inicial
    normalized = re.sub(r"[\s\-\.]", "", s)
    return normalized if normalized else None


def clean_curp(value) -> str | None:
    """Valida que el CURP tenga 18 caracteres alfanuméricos."""
    s = clean(value)
    if not s:
        return None
    s = s.upper().strip()
    if len(s) == 18 and re.match(r"^[A-Z0-9]{18}$", s):
        return s
    # Si tiene longitud distinta, igual lo guardamos pero advertimos
    return s


def clean_cp(value) -> str | None:
    """Código postal: conservar como texto (puede tener cero inicial)."""
    s = clean(value)
    if not s:
        return None
    # Quitar decimales si vino como float (ej. "7580.0")
    s = re.sub(r"\.0+$", "", s)
    return s.zfill(5) if s.isdigit() and len(s) < 5 else s


# ---------------------------------------------------------------------------
# Carga del Excel
# ---------------------------------------------------------------------------

def load_excel() -> list[dict]:
    """Lee la hoja CONSOLIDADO y devuelve lista de dicts por atleta."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("❌  Falta dependencia: pip3 install openpyxl")

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"❌  Hoja '{SHEET_NAME}' no encontrada en {EXCEL_PATH}")

    ws = wb[SHEET_NAME]
    athletes = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # Ignorar filas completamente vacías
        if all(v is None for v in row):
            continue

        folio = clean(row[COL_FOLIO])
        if not folio:
            print(f"  ⚠️  Fila {i+2}: sin Folio — omitida")
            continue

        curp    = clean_curp(row[COL_CURP])
        cp      = clean_cp(row[COL_CP])
        colonia = clean(row[COL_COLONIA])
        phone   = clean_phone(row[COL_TELEFONO])
        nombre  = clean(row[COL_NOMBRE])
        apellido = clean(row[COL_APELLIDO])

        athletes.append({
            "folio":    folio,
            "nombre":   f"{nombre} {apellido}".strip(),
            "curp":     curp,
            "cp":       cp,
            "colonia":  colonia,
            "phone":    phone,
        })

    return athletes


# ---------------------------------------------------------------------------
# Carga de atletas desde BD
# ---------------------------------------------------------------------------

def load_db_athletes() -> dict[str, dict]:
    """Devuelve dict keyed by athlete_code con id y campos actuales."""
    rows = api_get(
        "/rest/v1/athletes"
        "?select=id,athlete_code,first_name,last_name,curp,cp,colonia,phone"
        "&athlete_code=not.is.null"
        "&order=athlete_code.asc"
    )
    return {r["athlete_code"]: r for r in rows}


# ---------------------------------------------------------------------------
# Comparación y actualización
# ---------------------------------------------------------------------------

def needs_update(db_row: dict, excel_row: dict) -> dict | None:
    """Devuelve dict con los campos que cambiarían, o None si no hay cambio."""
    changes: dict = {}
    for field in ("curp", "cp", "colonia", "phone"):
        excel_val = excel_row.get(field)
        db_val    = db_row.get(field)

        # Solo actualizar si el Excel tiene datos y la BD no los tiene (o son distintos)
        if excel_val and excel_val != db_val:
            changes[field] = excel_val

    return changes if changes else None


def run_import(excel_athletes: list[dict], db_athletes: dict[str, dict]) -> None:
    updated    = 0
    skipped    = 0
    not_found  = []
    no_changes = 0

    for row in excel_athletes:
        folio = row["folio"]
        db    = db_athletes.get(folio)

        if not db:
            not_found.append(row)
            continue

        changes = needs_update(db, row)
        if not changes:
            no_changes += 1
            continue

        label = f"{folio} — {row['nombre']}"

        if DRY_RUN:
            print(f"  🔍 [{label}]  actualizaría: {changes}")
            skipped += 1
        else:
            try:
                api_patch(
                    f"/rest/v1/athletes?id=eq.{db['id']}",
                    changes,
                )
                print(f"  ✅ [{label}]  actualizado: {list(changes.keys())}")
                updated += 1
            except Exception as e:
                print(f"  ❌ [{label}]  error: {e}")

    # Resumen
    print()
    print("=" * 60)
    if DRY_RUN:
        print(f"DRY-RUN — no se escribió nada en la BD")
        print(f"  Actualizarían:  {skipped}")
    else:
        print(f"  Actualizados:   {updated}")
    print(f"  Sin cambios:    {no_changes}")
    print(f"  No encontrados: {len(not_found)}")

    if not_found and (SHOW_MISSING or DRY_RUN):
        print()
        print("Atletas en Excel sin coincidencia en BD:")
        for r in not_found:
            print(f"  {r['folio']:8}  {r['nombre']}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    global SUPABASE_URL, SUPABASE_KEY

    print("=" * 60)
    print("import_athlete_demographics.py")
    print(f"Modo: {'DRY-RUN (sin escritura)' if DRY_RUN else 'PRODUCCIÓN'}")
    print(f"Fuente: {Path(EXCEL_PATH).name}  /  hoja '{SHEET_NAME}'")
    print("=" * 60)
    print()

    SUPABASE_URL, SUPABASE_KEY = load_env()

    print("📋  Leyendo Excel…")
    excel_athletes = load_excel()
    print(f"    {len(excel_athletes)} atletas encontrados en Excel")

    print("🗄️   Leyendo BD…")
    db_athletes = load_db_athletes()
    print(f"    {len(db_athletes)} atletas con Folio en BD")
    print()

    run_import(excel_athletes, db_athletes)


if __name__ == "__main__":
    main()
